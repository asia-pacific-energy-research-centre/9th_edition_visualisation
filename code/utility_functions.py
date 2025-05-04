#%%
import time    
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import pandas as pd 
import re
import os
from datetime import datetime
import pickle
import shutil
from zipfile import BadZipFile
STRICT_DATA_CHECKING = False

# ECONOMY_ID = '22_SEA'#'00_APEC'#08_JPN#tremoved this because its just confusing. might as well set it in the script

#######################################################
#CONFIG PREPARATION
#create FILE_DATE_ID for use in file names
FILE_DATE_ID =datetime.now().strftime('%Y%m%d')#'20250401'#datetime.now().strftime('%Y%m%d')#'20250317'# '20250409'#'20250410'#'20250415'#

# FILE_DATE_ID = '20241112'
total_plotting_names=['Total', 'TPES', 'Total primary energy supply','TFEC', 'TFC', 'Total_industry', 'Total_transport', 'Total_fuels', 'Total combustion emissions', 'Total Industry & Non-energy', 'Total production', 'Total power fuel consumption', 'Total generation', 'Total generation capacity','Total use of fuels', 'TFC_hydrogen','TFC_low_carbon_fuels', 'TPES_bioenergy', 'TFC_refined_fuels', 'Refined products', 'TFEC_incl_own_use_losses', 'Refined products and low carbon fuels', 'TFC_refined_products_and_low_carbon_fuels']#dont add ,'Net emissions' to this!

EXPECTED_COLS = ['source', 'table_number', 'chart_type','plotting_name', 'plotting_name_column','aggregate_name', 'aggregate_name_column', 'scenario', 'unit', 'table_id', 'dimensions', 'chart_title', 'year', 'value','sheet_name']

SCENARIOS_list = ['reference', 'target']

ALL_ECONOMY_IDS = ["01_AUS", "02_BD", "03_CDA", "04_CHL", "05_PRC", "06_HKC", "07_INA", "08_JPN", "09_ROK", "10_MAS", "11_MEX", "12_NZ", "13_PNG", "14_PE", "15_PHL", "16_RUS", "17_SGP", "18_CT", "19_THA", "20_USA", "21_VN"]

AGGREGATE_ECONOMY_MAPPING = {
    '00_APEC': ['01_AUS', '02_BD', '03_CDA', '04_CHL', '05_PRC', '06_HKC', '07_INA', '08_JPN', '09_ROK', '10_MAS', '11_MEX', '12_NZ', '13_PNG', '14_PE', '15_PHL', '16_RUS', '17_SGP', '18_CT', '19_THA', '20_USA', '21_VN'],
    '22_SEA': ['02_BD', '07_INA', '10_MAS', '15_PHL', '17_SGP', '19_THA', '21_VN'],
    '23_NEA': ['06_HKC', '08_JPN', '09_ROK', '18_CT'],
    '23b_ONEA': ['06_HKC', '09_ROK', '18_CT'],
    '24_OAM': ['03_CDA', '04_CHL', '11_MEX', '14_PE'],
    '24b_OOAM': ['04_CHL', '11_MEX', '14_PE'],
    '25_OCE': ['01_AUS', '12_NZ', '13_PNG'],
    '26_NA': ['03_CDA', '20_USA'],
}

MIN_YEAR = 2000
EBT_EARLIEST_YEAR = 1980
OUTLOOK_BASE_YEAR = 2022 #if running vis for russia, set this to the same as OUTLOOK_BASE_YEAR_RUSSIA
OUTLOOK_BASE_YEAR_RUSSIA = 2021
OUTLOOK_LAST_YEAR = 2060

def check_base_year_is_as_expected(ECONOMY_ID):
    #check that the OUTLOOK_BASE_YEAR in utils is == to OUTLOOK_BASE_YEAR_russia in utils, else raise an error
    #and if SINGLE_ECONOMY_ID != '16_RUS' then check that OUTLOOK_BASE_YEAR in utils != OUTLOOK_BASE_YEAR_russia in utils, else raise an error
    if OUTLOOK_BASE_YEAR != OUTLOOK_BASE_YEAR_RUSSIA and ECONOMY_ID == '16_RUS':
        raise ValueError('OUTLOOK_BASE_YEAR in utility_functions is not equal to OUTLOOK_BASE_YEAR_RUSSIA in utility_functions')
    elif ECONOMY_ID != '16_RUS':
        if OUTLOOK_BASE_YEAR == OUTLOOK_BASE_YEAR_RUSSIA:
            raise ValueError('OUTLOOK_BASE_YEAR in utility_functions is equal to OUTLOOK_BASE_YEAR_RUSSIA in utility_functions')
        
PROJ_YEARS = list(range(OUTLOOK_BASE_YEAR+1, OUTLOOK_LAST_YEAR+1, 1))
HIST_YEARS = list(range(EBT_EARLIEST_YEAR, OUTLOOK_BASE_YEAR+1, 1))
HIST_YEARS_str = [str(year) for year in HIST_YEARS]
PROJ_YEARS_str = [str(year) for year in PROJ_YEARS]

def find_most_recent_file_date_id(directory_path, filename_part = None,RETURN_DATE_ID = False, PRINT=True):
    """Find the most recent file in a directory based on the date ID in the filename."""
    # List all files in the directory
    files = os.listdir(directory_path)

    # Initialize variables to keep track of the most recent file and date
    most_recent_date = datetime.min
    most_recent_file = None
    date_id = None
    # Define a regex pattern for the date ID (format YYYYMMDD)
    date_pattern = re.compile(r'(\d{8})')
    
    # Loop through the files to find the most recent one
    for file in files:
        if filename_part is not None:
            if filename_part not in file:
                continue
        # Use regex search to find the date ID in the filename
        if os.path.isdir(os.path.join(directory_path, file)):
            continue
        match = date_pattern.search(file)
        if match:
            date_id = match.group(1)
            # Parse the date ID into a datetime object
            try:
                file_date = datetime.strptime(date_id, '%Y%m%d')
                # If this file's date is more recent, update the most recent variables
                if file_date > most_recent_date:
                    most_recent_date = file_date
                    most_recent_file = file
            except ValueError:
                # If the date ID is not in the expected format, skip this file
                continue

    # Output the most recent file
    if most_recent_file:
        if PRINT:
            print(f"The most recent file is: {most_recent_file} with the date ID {most_recent_date.strftime('%Y%m%d')}")
    else:
        if PRINT:
            print("No files found with a valid date ID.")
    if RETURN_DATE_ID:
        return most_recent_file, date_id
    else:
        return most_recent_file

def import_files_from_ebt_system(ECONOMY_ID, ebt_system_file_path='../../Outlook9th_EBT/results/'):
    print('Automatically importing files from EBT system')
    emissions_no2 = ebt_system_file_path + f'{ECONOMY_ID}/emissions/emissions_no2_{ECONOMY_ID}_DDDDDDDD.csv'
    emissions_co2 = ebt_system_file_path + f'{ECONOMY_ID}/emissions/emissions_co2_{ECONOMY_ID}_DDDDDDDD.csv'
    emissions_ch4 = ebt_system_file_path + f'{ECONOMY_ID}/emissions/emissions_ch4_{ECONOMY_ID}_DDDDDDDD.csv'
    emissions_co2e = ebt_system_file_path + f'{ECONOMY_ID}/emissions/emissions_co2e_{ECONOMY_ID}_DDDDDDDD.csv'
    energy_folder = ebt_system_file_path + f'{ECONOMY_ID}/merged/merged_file_energy_{ECONOMY_ID}_DDDDDDDD.csv'
    capacity_folder = ebt_system_file_path + f'{ECONOMY_ID}/capacity/capacity_{ECONOMY_ID}_DDDDDDDD.csv'
    
    old_files = [file for file in os.listdir(f'../input_data/{ECONOMY_ID}') if file.endswith('.csv')]
    files = [emissions_no2, emissions_co2, emissions_ch4, emissions_co2e, energy_folder, capacity_folder]
    files_with_dates = []
    filenames_with_dates = []
    #search for file with the latest date to replace DDDDDDDD
    for file in files:
        directory_path = file.split('/')[:-1]
        directory_path = '/'.join(directory_path)
        filename_part = file.split('/')[-1].strip('_DDDDDDDD.csv')
        try:
            filename, date_id = find_most_recent_file_date_id(directory_path, filename_part, RETURN_DATE_ID = True, PRINT=False)
        except:
            breakpoint()
            raise Exception(f'No files found in {directory_path} with the pattern {filename_part}')
        file = file.replace('DDDDDDDD', date_id)
        files_with_dates.append(file)
        filenames_with_dates.append(filename)
        
    #now move the file to the correct location and move the old one to archive
    for old_file in old_files:
        if not os.path.exists(f'../input_data/{ECONOMY_ID}/archive/'):
            os.mkdir(f'../input_data/{ECONOMY_ID}/archive/')
        shutil.move(f'../input_data/{ECONOMY_ID}/{old_file}', f'../input_data/{ECONOMY_ID}/archive/{old_file}') 
        
    for file, filename in zip(files_with_dates, filenames_with_dates):  
        shutil.copy(file, f'../input_data/{ECONOMY_ID}/{filename}')
        
        
def save_checkpoint(df, name, folder='../intermediate_data/checkpoints/'):
    """
    Save a DataFrame or data object as a checkpoint using pickle.
    """
    if not os.path.exists(folder):
        os.makedirs(folder)
    filepath = os.path.join(folder, f"{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pkl")
    with open(filepath, 'wb') as f:
        try:
            pickle.dump(df, f)
            print(f"Checkpoint saved: {filepath}")
        except TypeError as e:
            print(f"Failed to save checkpoint {name}: {e}")

def load_checkpoint(name, folder='../intermediate_data/checkpoints/'):
    """
    Load a checkpoint from a pickle file.
    """
    checkpoints = [f for f in os.listdir(folder) if name in f and f.endswith('.pkl')]
    if not checkpoints:
        raise FileNotFoundError(f"No checkpoint found for {name}")
    checkpoints.sort()  # Sort by timestamp
    filepath = os.path.join(folder, checkpoints[-1])  # Load the latest checkpoint
    with open(filepath, 'rb') as f:
        data = pickle.load(f)
    print(f"Checkpoint loaded: {filepath}")
    return data

def data_checking_warning_or_error(message):
    if STRICT_DATA_CHECKING:
        raise Exception(message)
    else:
        print(message)

def move_workbooks_to_onedrive(origin_date_id=FILE_DATE_ID, econ_list=ALL_ECONOMY_IDS):
    # clean_onedrive_workbooks_folder(date_ids_to_keep='20241211', specific_files_to_keep=[], econ_list=ALL_ECONOMY_IDS, archive_folder_name="first_iteration")
    # clean_onedrive_workbooks_folder(date_ids_to_keep='20241211', specific_files_to_keep=[], econ_list=AGGREGATE_ECONOMY_MAPPING.keys(), archive_folder_name="first_iteration")
    CURRENT_DATE_ID = datetime.now().strftime("%Y%m%d")
    for economy_id in econ_list:
        source_path = f'C:/Users/finbar.maunsell/github/9th_edition_visualisation/output/output_workbooks/{economy_id}/{economy_id}_charts_{origin_date_id}.xlsx'
        destination_path = f'C:/Users/finbar.maunsell/OneDrive - APERC/outlook 9th/Modelling/Visualisation/{economy_id}/{economy_id}_charts_{CURRENT_DATE_ID}.xlsx'
        
        # Create destination directory if it doesn't exist
        os.makedirs(os.path.dirname(destination_path), exist_ok=True)
        
        # Move the file
        shutil.copy(source_path, destination_path)
        print(f"Moved {source_path} to {destination_path}")

def clean_onedrive_workbooks_folder(date_ids_to_keep, specific_files_to_keep, econ_list, archive_folder_name="archive"):
    # clean_onedrive_workbooks_folder(date_ids_to_keep=['20241211'], specific_files_to_keep=[], econ_list=["01_AUS", "02_BD", "03_CDA", "04_CHL", "05_PRC", "06_HKC", "07_INA", "08_JPN", "09_ROK", "10_MAS", "11_MEX", "12_NZ", "13_PNG", "14_PE", "15_PHL", "16_RUS", "17_SGP", "18_CT", "19_THA", "20_USA", "21_VN"], archive_folder_name="first_iteration")
    # clean_onedrive_workbooks_folder(date_ids_to_keep=['20241211'], specific_files_to_keep=[], econ_list=AGGREGATE_ECONOMY_MAPPING.keys(), archive_folder_name="first_iteration")
    for economy_id in econ_list:
        base_path = f'C:/Users/finbar.maunsell/OneDrive - APERC/outlook 9th/Modelling/Visualisation/{economy_id}/'
        archive_path = os.path.join(base_path, archive_folder_name)
        
        # Create archive directory if it doesn't exist
        os.makedirs(archive_path, exist_ok=True)
        
        for file_name in os.listdir(base_path):
            file_path = os.path.join(base_path, file_name)
            
            # Check if it's a file and not a directory
            if os.path.isfile(file_path):
                # Check if the file does not contain any of the date_ids_to_keep and is not in specific_files_to_keep
                if all(date_id not in file_name for date_id in date_ids_to_keep) and file_name not in specific_files_to_keep:
                    shutil.move(file_path, os.path.join(archive_path, file_name))
                    print(f"Moved {file_path} to {archive_path}")
          
def highlight_differences_in_master_config_xlsx(file1, file2, output_file):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    import pandas as pd
    import numpy as np

    # Load both workbooks
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)

    # Create a new workbook for output
    output_wb = Workbook()
    output_wb.remove(output_wb.active)  # Remove the default sheet

    fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill

    for sheet_name in wb1.sheetnames:
        if sheet_name in wb2.sheetnames:
            # Read both sheets into DataFrames
            df1 = pd.read_excel(file1, sheet_name=sheet_name)
            df2 = pd.read_excel(file2, sheet_name=sheet_name)

            # Compare the sheets
            # if sheet_name == 'two_dimensional_plots':
            #     breakpoint()
            differences = compare_sheets(df1, df2)

            if differences:  # Only create a sheet if there are differences
                output_ws = output_wb.create_sheet(sheet_name)

                # Copy the DataFrame to the output sheet, ensuring all columns are present
                for r_idx, row_data in enumerate(df1.itertuples(index=False), start=1):
                    output_ws.append(row_data)

                # Highlight differences
                for diff in differences:
                    if len(diff) == 3:  # Handle missing or new rows
                        row, message, values = diff
                        output_ws.append([message] + values)  # Add message and row data to the output sheet
                        cell = output_ws.cell(row=row, column=1)  # Ensure the row starts from the correct position
                        cell.fill = fill_color

                    elif len(diff) == 4:  # Handle individual cell differences
                        row, col, val1, val2 = diff
                        cell = output_ws.cell(row=row + 1, column=col + 1)  # Adjust for 1-based Excel indexing
                        cell.fill = fill_color
                        cell.comment = Comment(f"Old: {val1}, New: {val2}", "Comparison Script")

    # Save the output file only if any sheets were created
    if output_wb.sheetnames:
        output_wb.save(output_file)
        print(f"Comparison complete. Results saved to '{output_file}'.")
    else:
        print("No differences found. No file was created.")


def compare_sheets(sheet1, sheet2):
    import numpy as np

    differences = []
    
    # Ensure both sheets have the same shape and fill missing rows/columns with NaN
    max_rows = max(sheet1.shape[0], sheet2.shape[0])
    max_cols = max(sheet1.shape[1], sheet2.shape[1])
    
    #get cols by index number
    cols_to_keep_sheet1 = sheet1.columns.tolist()[0:max_cols]
    cols_to_keep_sheet2 = sheet2.columns.tolist()[0:max_cols]

    sheet1 = sheet1.reindex(index=range(max_rows), columns=cols_to_keep_sheet1, fill_value=np.nan)
    sheet2 = sheet2.reindex(index=range(max_rows), columns=cols_to_keep_sheet2, fill_value=np.nan)
    
    # Detect and flag missing or new rows
    for row in range(max_rows):
        row1 = sheet1.iloc[row].dropna().tolist()
        row2 = sheet2.iloc[row].dropna().tolist()

        if not row1 and row2:
            differences.append((row + 1, "MISSING IN FILE 1", row2))
        elif row1 and not row2:
            differences.append((row + 1, "MISSING IN FILE 2", row1))

    # Compare each cell for remaining rows
    for row in range(max_rows):
        for col in range(max_cols):
            val1 = sheet1.iloc[row, col]
            val2 = sheet2.iloc[row, col]

            # Check for NaN equality
            if pd.isna(val1) and pd.isna(val2):
                continue  # Skip if both values are NaN

            # Ensure both values have the same type for accurate comparison
            if isinstance(val1, (int, float)) and isinstance(val2, (int, float)):
                # Allow small floating-point differences due to precision issues
                if abs(val1 - val2) < 1e-9:
                    continue

            # Add cell differences to the list
            if val1 != val2:
                differences.append((row, col, val1, val2))  # Adjust for 1-based indexing

    return differences

def clean_up_old_files(ECONOMY_ID):
    #empty this folder ifit exists:
    folder_path = f'../output/plotting_output/{ECONOMY_ID}'
    
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)
    os.makedirs(folder_path)
    # intermediate_data\charts
    folder_path = '../intermediate_data/charts'
    
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)
    os.makedirs(folder_path)
    
def save_used_colors_dict(colors_dict):
    #this is used in check_plotting_names_in_colours_dict()
    start_time = time.time()
    
    #for helping to remove plotting names that arent used anymore, keep a record of all plotting names that are actually used. this will get cleaned every once in a while
    
    #open up the csv and where there arent any matches, enter into the bottom. have a colomn for plotting_name and a column for color
    used_colors_excel_path = '../config/used_colors_dict.xlsx'

    # Load the Excel file or create a new one if it doesn't exist
    try:
        workbook = load_workbook(used_colors_excel_path)
        sheet = workbook.active
    except Exception as e:  # Handle multiple exceptions
        if isinstance(e, BadZipFile):
            print(f"Skipped processing {used_colors_excel_path} as it might be open in another script.")
            return  # Skip further processing
        breakpoint()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Used Colors"
        sheet.append(["plotting_name", "color"])  # Add headers
        print(f"Exception occurred: {e}. Created a new workbook.")

    # Create a dictionary from the existing data in the Excel file
    used_colors_dict = {row[0]: row[1] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]}

    # Add new plotting names and colors to the Excel file
    for plotting_name, color in colors_dict.items():
        if plotting_name not in used_colors_dict:
            sheet.append([plotting_name, color])
            used_colors_dict[plotting_name] = color

    # Apply color fill to the color cells
    for row in sheet.iter_rows(min_row=2, max_col=2):
        color_cell = row[1]
        if color_cell.value and isinstance(color_cell.value, str) and color_cell.value.startswith("#"):
            try:
                color_fill = PatternFill(start_color=color_cell.value[1:], end_color=color_cell.value[1:], fill_type="solid")
                color_cell.fill = color_fill
            except ValueError:
                print(f"Invalid color value: {color_cell.value}")

    # Save the Excel file
    workbook.save(used_colors_excel_path)
    # print(f"Used colors saved to {used_colors_excel_path}")
    
    end_time = time.time()
    # print(f"Execution time: {end_time - start_time:.2f} seconds")
#%%
# if __name__ == "__main__":
#     # RUN THIS FILE IF YOU WANT TO RUN THE BELOW:
    
#     file1_path = "../config/master_config.xlsx"
#     #find other file in the folder that has master config in the name. if there is no other one or multiple, raise an error
#     files = os.listdir("../config")
#     files = [f for f in files if 'master_config' in f]
#     #ignore any files with ~$ in the name since they are temporary files
#     files = [f for f in files if '~$' not in f]
#     if len(files) != 2:
#         raise ValueError("There are not exactly 2 files with 'master_config' in the name in the config folder.")
#     file2_path = os.path.join("../config", [f for f in files if f != 'master_config.xlsx'][0])
#     output_file_path = "../comparison_master_config.xlsx"

# highlight_differences_in_master_config_xlsx(file1_path, file2_path, output_file_path)
    
#%%

def move_workbooks_to_onedrive(origin_date_id=FILE_DATE_ID, econ_list=ALL_ECONOMY_IDS, FIND_LATEST_FILE_ID=True):
    # clean_onedrive_workbooks_folder(date_ids_to_keep='20241211', specific_files_to_keep=[], econ_list=ALL_ECONOMY_IDS, archive_folder_name="first_iteration")
    # clean_onedrive_workbooks_folder(date_ids_to_keep='20241211', specific_files_to_keep=[], econ_list=AGGREGATE_ECONOMY_MAPPING.keys(), archive_folder_name="first_iteration")
    # CURRENT_DATE_ID = datetime.now().strftime("%Y%m%d")
    for economy_id in econ_list:
        source_folder = f'C:/Users/finbar.maunsell/github/9th_edition_visualisation/output/output_workbooks/{economy_id}/'
        source_file = f'{economy_id}_charts_{origin_date_id}.xlsx'
        source_path = f'{source_folder}/{source_file}'
        destination_folder = f'C:/Users/finbar.maunsell/OneDrive - APERC/outlook 9th/Modelling/Visualisation/{economy_id}'
        destination_file = f'{economy_id}_charts_{origin_date_id}.xlsx'
        destination_path = f'{destination_folder}/{destination_file}'
        
        if FIND_LATEST_FILE_ID and not os.path.exists(
            source_path
        ):
            f, latest_file_id = find_most_recent_file_date_id(
                source_folder, RETURN_DATE_ID=True
            )
            if latest_file_id != None:
                source_path = source_path.replace(origin_date_id, latest_file_id)
                destination_path = destination_path.replace(origin_date_id, latest_file_id)
            else:
                raise ValueError("Latest file ID is None, cannot proceed.")
            
        # Create destination directory if it doesn't exist
        os.makedirs(os.path.dirname(destination_path), exist_ok=True)
        
        # Move the file
        shutil.copy(source_path, destination_path)
        print(f"Moved {source_path} to {destination_path}")
        
# #%%
# move_workbooks_to_onedrive(origin_date_id=FILE_DATE_ID, econ_list= [ '14_PE', '20_USA', '08_JPN', '12_NZ', '17_SGP', '03_CDA','13_PNG'])

#%%